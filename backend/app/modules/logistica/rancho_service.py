from __future__ import annotations

from calendar import monthrange
from datetime import date
from io import BytesIO

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, selectinload

from app.models.police_officer import PoliceOfficer
from app.models.rancho_configuracao import RanchoConfiguracao
from app.models.rancho_lancamento import RanchoLancamento
from app.models.rancho_participante import RanchoParticipante
from app.models.user import User
from app.shared.utils.scope import (
    apply_unit_scope,
    can_access_unit,
    resolve_filter_unit_ids,
    resolve_unit_id_for_creation,
)

WEEKDAY_SHORT = ["seg.", "ter.", "qua.", "qui.", "sex.", "sáb.", "dom."]
MONTH_SHORT = {
    1: "jan.",
    2: "fev.",
    3: "mar.",
    4: "abr.",
    5: "mai.",
    6: "jun.",
    7: "jul.",
    8: "ago.",
    9: "set.",
    10: "out.",
    11: "nov.",
    12: "dez.",
}


def _base_query(db: Session, current_user: User):
    return (
        apply_unit_scope(db.query(RanchoConfiguracao), RanchoConfiguracao, current_user)
        .options(
            joinedload(RanchoConfiguracao.unit),
            joinedload(RanchoConfiguracao.criado_por),
            selectinload(RanchoConfiguracao.participantes).selectinload(RanchoParticipante.lancamentos),
        )
    )


def _get_config_or_404(db: Session, current_user: User, config_id: int) -> RanchoConfiguracao:
    config = _base_query(db, current_user).filter(RanchoConfiguracao.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Planejamento de rancho não encontrado.")
    return config


def _require_admin_to_close(current_user: User) -> None:
    if current_user.role_code not in {"ADMIN_GLOBAL", "ADMIN_UNIDADE"} and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem fechar o planejamento.")


def _normalize_re(value: str | None) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _month_dates(mes: int, ano: int) -> list[date]:
    return [
        current
        for current in (date(ano, mes, day) for day in range(1, monthrange(ano, mes)[1] + 1))
        if current.weekday() < 5
    ]


def _is_weekend(value: date) -> bool:
    return value.weekday() >= 5


def _resolve_pm_for_config(db: Session, config: RanchoConfiguracao, re_value: str) -> PoliceOfficer:
    normalized_re = _normalize_re(re_value)
    if not normalized_re:
        raise HTTPException(status_code=400, detail="Informe um RE válido.")

    try:
        allowed_unit_ids = resolve_filter_unit_ids(db, config.unit_id)
    except ValueError:
        allowed_unit_ids = {config.unit_id}

    officers = (
        db.query(PoliceOfficer)
        .options(joinedload(PoliceOfficer.unit))
        .filter(PoliceOfficer.unit_id.in_(allowed_unit_ids))
        .all()
    )
    officer = next((item for item in officers if _normalize_re(item.re_with_digit) == normalized_re), None)
    if not officer:
        raise HTTPException(status_code=404, detail="RE não encontrado no cadastro de policiais.")
    return officer


def _participant_to_dict(participante: RanchoParticipante) -> dict:
    return {
        "id": participante.id,
        "configuracao_id": participante.configuracao_id,
        "tipo_pessoa": participante.tipo_pessoa,
        "re": participante.re,
        "rg": participante.rg,
        "nome": participante.nome,
        "graduacao": participante.graduacao,
        "ordem": participante.ordem,
        "display_name": participante.display_name,
        "lancamentos": [
            {
                "id": lancamento.id,
                "participante_id": lancamento.participante_id,
                "data": lancamento.data,
                "cafe": lancamento.cafe,
                "almoco": lancamento.almoco,
            }
            for lancamento in sorted(participante.lancamentos, key=lambda item: item.data)
        ],
    }


def _build_resumo_for_type(participantes: list[RanchoParticipante], dates: list[date], tipo_pessoa: str | None):
    selected = [item for item in participantes if tipo_pessoa is None or item.tipo_pessoa == tipo_pessoa]
    resumo: list[dict] = []
    for current_date in dates:
        total_cafe = 0
        total_almoco = 0
        for participante in selected:
            lancamento = next((item for item in participante.lancamentos if item.data == current_date), None)
            if not lancamento:
                continue
            total_cafe += 1 if lancamento.cafe else 0
            total_almoco += 1 if lancamento.almoco else 0
        resumo.append(
            {
                "data": current_date,
                "total_cafe": total_cafe,
                "total_almoco": total_almoco,
                "total_geral": total_cafe + total_almoco,
            }
        )
    return resumo


def _config_to_detail(config: RanchoConfiguracao) -> dict:
    dates = _month_dates(config.mes, config.ano)
    participantes = sorted(config.participantes, key=lambda item: (item.ordem, item.id))
    return {
        "id": config.id,
        "mes": config.mes,
        "ano": config.ano,
        "unit_id": config.unit_id,
        "unit_label": config.unit_label,
        "criado_por_id": config.criado_por_id,
        "criado_por_nome": config.criado_por_nome,
        "criado_em": config.criado_em,
        "fechado": config.fechado,
        "participantes": [_participant_to_dict(item) for item in participantes],
        "totais_pm": _build_resumo_for_type(participantes, dates, "PM"),
        "totais_civil": _build_resumo_for_type(participantes, dates, "CIVIL"),
        "totais_visitante": _build_resumo_for_type(participantes, dates, "VISITANTE"),
        "totais_geral": _build_resumo_for_type(participantes, dates, None),
    }


def list_configuracoes(*, mes: int | None, ano: int | None, unidade_id: int | None, db: Session, current_user: User):
    query = _base_query(db, current_user)
    if mes is not None:
        query = query.filter(RanchoConfiguracao.mes == mes)
    if ano is not None:
        query = query.filter(RanchoConfiguracao.ano == ano)
    if unidade_id is not None:
        if not can_access_unit(current_user, unidade_id):
            raise HTTPException(status_code=403, detail="Sem permissão para a unidade informada.")
        try:
            allowed_ids = resolve_filter_unit_ids(db, unidade_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Unidade informada não encontrada.")
        query = query.filter(RanchoConfiguracao.unit_id.in_(allowed_ids))

    items = query.order_by(RanchoConfiguracao.ano.desc(), RanchoConfiguracao.mes.desc(), RanchoConfiguracao.criado_em.desc()).all()
    return [
        {
            "id": item.id,
            "mes": item.mes,
            "ano": item.ano,
            "unit_id": item.unit_id,
            "unit_label": item.unit_label,
            "criado_por_id": item.criado_por_id,
            "criado_por_nome": item.criado_por_nome,
            "criado_em": item.criado_em,
            "fechado": item.fechado,
            "total_participantes": len(item.participantes),
        }
        for item in items
    ]


def create_configuracao(*, payload, db: Session, current_user: User):
    unidade_id = resolve_unit_id_for_creation(current_user, payload.unidade_id)
    existing = (
        _base_query(db, current_user)
        .filter(
            RanchoConfiguracao.unit_id == unidade_id,
            RanchoConfiguracao.mes == payload.mes,
            RanchoConfiguracao.ano == payload.ano,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Já existe planejamento para esta unidade neste mês/ano.")

    item = RanchoConfiguracao(
        mes=payload.mes,
        ano=payload.ano,
        unit_id=unidade_id,
        criado_por_id=current_user.id,
        fechado=False,
    )
    db.add(item)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Já existe planejamento para esta unidade neste mês/ano.")
    return _config_to_detail(_get_config_or_404(db, current_user, item.id))


def get_configuracao(*, config_id: int, db: Session, current_user: User):
    return _config_to_detail(_get_config_or_404(db, current_user, config_id))


def buscar_pm(*, re: str, unidade_id: int | None, db: Session, current_user: User):
    target_unit_id = unidade_id if unidade_id and can_access_unit(current_user, unidade_id) else current_user.unit_id
    dummy_config = RanchoConfiguracao(unit_id=target_unit_id, mes=1, ano=2000, criado_por_id=current_user.id)
    officer = _resolve_pm_for_config(db, dummy_config, re)
    return {
        "policial_id": officer.id,
        "re": officer.re_with_digit,
        "nome": officer.war_name or officer.full_name,
        "nome_completo": officer.full_name,
        "graduacao": officer.rank,
        "unidade": officer.unit_label,
    }


def add_participante(*, config_id: int, payload, db: Session, current_user: User):
    config = _get_config_or_404(db, current_user, config_id)
    ordem = payload.ordem if payload.ordem is not None else max((item.ordem for item in config.participantes), default=0) + 1

    if payload.tipo_pessoa == "PM":
        officer = _resolve_pm_for_config(db, config, payload.re or "")
        participante = RanchoParticipante(
            configuracao_id=config.id,
            tipo_pessoa="PM",
            re=officer.re_with_digit,
            rg=None,
            nome=officer.war_name or officer.full_name,
            graduacao=officer.rank,
            ordem=ordem,
        )
    else:
        participante = RanchoParticipante(
            configuracao_id=config.id,
            tipo_pessoa=payload.tipo_pessoa,
            re=None,
            rg=payload.rg if payload.tipo_pessoa == "CIVIL" else None,
            nome=payload.nome or "",
            graduacao=payload.graduacao,
            ordem=ordem,
        )

    db.add(participante)
    db.commit()
    return _config_to_detail(_get_config_or_404(db, current_user, config.id))


def remove_participante(*, config_id: int, participante_id: int, db: Session, current_user: User):
    config = _get_config_or_404(db, current_user, config_id)
    participante = next((item for item in config.participantes if item.id == participante_id), None)
    if not participante:
        raise HTTPException(status_code=404, detail="Participante não encontrado no planejamento.")
    db.delete(participante)
    db.commit()
    return _config_to_detail(_get_config_or_404(db, current_user, config.id))


def upsert_lancamento(*, config_id: int, payload, db: Session, current_user: User):
    config = _get_config_or_404(db, current_user, config_id)
    participante = next((item for item in config.participantes if item.id == payload.participante_id), None)
    if not participante:
        raise HTTPException(status_code=404, detail="Participante não encontrado no planejamento.")
    if payload.data.month != config.mes or payload.data.year != config.ano:
        raise HTTPException(status_code=400, detail="A data informada está fora do mês/ano do planejamento.")
    if _is_weekend(payload.data):
        raise HTTPException(status_code=400, detail="Não é permitido lançar rancho em sábados ou domingos.")

    lancamento = (
        db.query(RanchoLancamento)
        .filter(RanchoLancamento.participante_id == participante.id, RanchoLancamento.data == payload.data)
        .first()
    )
    if lancamento:
        lancamento.cafe = payload.cafe
        lancamento.almoco = payload.almoco
    else:
        db.add(
            RanchoLancamento(
                participante_id=participante.id,
                data=payload.data,
                cafe=payload.cafe,
                almoco=payload.almoco,
            )
        )

    db.commit()
    return _config_to_detail(_get_config_or_404(db, current_user, config.id))


def fechar_configuracao(*, config_id: int, db: Session, current_user: User):
    _require_admin_to_close(current_user)
    config = _get_config_or_404(db, current_user, config_id)
    config.fechado = True
    db.commit()
    return _config_to_detail(_get_config_or_404(db, current_user, config.id))


def exportar_excel(*, config_id: int, db: Session, current_user: User):
    config = _get_config_or_404(db, current_user, config_id)
    detail = _config_to_detail(config)
    dates = _month_dates(config.mes, config.ano)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Previsão de Rancho"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="93B5DB")
    total_fill = PatternFill("solid", fgColor="4B5563")
    total_font = Font(color="FFFFFF", bold=True)
    marked_fill = PatternFill("solid", fgColor="C6EFCE")
    marked_font = Font(color="276221", bold=True)
    centered = Alignment(horizontal="center", vertical="center")

    last_col = 1 + (len(dates) * 2)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = sheet.cell(row=1, column=1, value=f"PREVISÃO DE RANCHO - {MONTH_SHORT[config.mes].upper()} {config.ano}")
    title_cell.fill = title_fill
    title_cell.font = Font(color="000000", bold=True, size=14)
    title_cell.alignment = centered

    sheet.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    name_header = sheet.cell(row=2, column=1, value="NOME")
    name_header.fill = header_fill
    name_header.font = header_font
    name_header.alignment = centered

    col = 2
    for current_date in dates:
        sheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        top = sheet.cell(row=2, column=col, value=f"{WEEKDAY_SHORT[current_date.weekday()]} {current_date.day}-{MONTH_SHORT[current_date.month]}")
        top.fill = header_fill
        top.font = header_font
        top.alignment = centered
        for idx, sub in enumerate(("C", "A")):
            cell = sheet.cell(row=3, column=col + idx, value=sub)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
        col += 2

    row_cursor = 4
    for participante in detail["participantes"]:
        sheet.cell(row=row_cursor, column=1, value=participante["display_name"])
        launches = {str(item["data"]): item for item in participante["lancamentos"]}
        col = 2
        for current_date in dates:
            current_launch = launches.get(str(current_date))
            for idx, key in enumerate(("cafe", "almoco")):
                marked = bool(current_launch and current_launch[key])
                cell = sheet.cell(row=row_cursor, column=col + idx, value="x" if marked else "")
                cell.alignment = centered
                if marked:
                    cell.fill = marked_fill
                    cell.font = marked_font
            col += 2
        row_cursor += 1

    for label, totals in [
        ("TOTAL EFETIVO PM", detail["totais_pm"]),
        ("FUNCIONÁRIAS CIVIS", detail["totais_civil"]),
        ("AVULSOS", detail["totais_visitante"]),
        ("TOTAL GERAL", detail["totais_geral"]),
    ]:
        cell = sheet.cell(row=row_cursor, column=1, value=label)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = centered
        col = 2
        for item in totals:
            for idx, value in enumerate((item["total_cafe"], item["total_almoco"])):
                total_cell = sheet.cell(row=row_cursor, column=col + idx, value=value)
                total_cell.fill = total_fill
                total_cell.font = total_font
                total_cell.alignment = centered
            col += 2
        row_cursor += 1

    for row in sheet.iter_rows(min_row=1, max_row=row_cursor - 1, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = border

    sheet.column_dimensions["A"].width = 34
    for col_index in range(2, last_col + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 4.2

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    file_name = f"previsao_rancho_{config.ano}_{str(config.mes).zfill(2)}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )
